#!/usr/bin/env python3
"""PLUR Enterprise — Organizational Learning + Knowledge Engineering + Datacore vision."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()

# --- Styles ---
header_font = Font(name="Calibri", size=14, bold=True)
section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subsection_font = Font(name="Calibri", size=11, bold=True)
subsection_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
body_font = Font(name="Calibri", size=11)
total_font = Font(name="Calibri", size=11, bold=True)
grand_font = Font(name="Calibri", size=12, bold=True)
discount_font = Font(name="Calibri", size=12, bold=True, color="2E7D32")
discount_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
grey_font = Font(name="Calibri", size=11, color="999999")
note_font = Font(name="Calibri", size=10, color="666666")
italic_grey = Font(name="Calibri", size=11, italic=True, color="666666")
recurring_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
recurring_font = Font(name="Calibri", size=12, bold=True, color="E65100")
param_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
param_font = Font(name="Calibri", size=11, bold=True, color="E65100")
annual_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
annual_font = Font(name="Calibri", size=13, bold=True, color="1B5E20")
margin_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
margin_font = Font(name="Calibri", size=12, bold=True, color="6A1B9A")
vision_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
datacore_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
datacore_font = Font(name="Calibri", size=12, bold=True, color="283593")
thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
thick_border = Border(top=Side(style="medium", color="2F5496"), bottom=Side(style="medium", color="2F5496"))
phase_border = Border(top=Side(style="thin", color="2F5496"), bottom=Side(style="thin", color="2F5496"))
eur_fmt = '#,##0 "EUR"'
NC = 8

def style_param(cell):
    cell.fill = param_fill
    cell.font = param_font
    cell.border = Border(
        left=Side(style="thin", color="E65100"), right=Side(style="thin", color="E65100"),
        top=Side(style="thin", color="E65100"), bottom=Side(style="thin", color="E65100"),
    )

team_info = [
    ("Gregor", "Project Director, Lead Dev"),
    ("Tadej", "CTO, Tech Lead"),
    ("Marko", "DevOps"),
    ("Crt", "PM, Data Scientist"),
]

# Subscription includes
subscription_includes = [
    "PLUR Enterprise server (HTTP/SSE MCP, multi-user)",
    "PostgreSQL + AGE (knowledge graph) + pgvector (semantic search)",
    "SSO integration (configured for your identity provider)",
    "Scope-based access control + role-level permissions",
    "MCP tool security (allowlist, write enforcement, audit)",
    "Admin dashboard (usage, health, audit log)",
    "Deployment, TLS, CI/CD, monitoring & alerting",
    "Security patches, dependency updates, platform upgrades",
    "Infrastructure hosting & AI compute (included)",
    "Priority bug fixes (< 24h response, < 72h resolution)",
    "Weekly check-in + quarterly strategic review",
]

# Custom services (Gregor, Tadej, Marko, Crt)
custom_phases = [
    {
        "name": "Integration \u2014 GitLab & Deployment",
        "token_cost": 200,
        "items": [
            ("GitLab OAuth2/OIDC + PKCE integration",                          0, 16,  0,  0),
            ("GitLab group/project membership sync & webhooks",                0, 12,  0,  0),
            ("Deployment & configuration for your GitLab instance",            0,  4,  8,  0),
            ("Setup workshop (all hands, 2-3h)",                               4,  0,  0,  4),
            ("Project management & weekly check-ins (across all phases)",       0,  0,  0, 16),
        ],
    },
    {
        "name": "Phase A \u2014 30 Active Repos (high-touch curation)",
        "token_cost": 600,
        "items": [
            ("Extraction strategy & planning with your tech leads",            0,  0,  0,  8),
            ("Convention extraction & analysis (30 repos, multiple passes)",   4,  0,  0, 24),
            ("Engram generation + quality tuning with your tech leads",         8,  0,  0, 16),
            ("Knowledge pack curation (per-project + per-group)",              4,  0,  0, 12),
            ("Custom ingest pipeline development (reusable for Phase B)",      0, 16,  0,  4),
            ("Phase report & deliverables review",                             0,  0,  0,  6),
        ],
    },
    {
        "name": "Phase B \u2014 1,370 Repos (automated processing)",
        "token_cost": 400,
        "items": [
            ("Batch processing \u2014 run pipeline on all remaining repos",    0,  0,  0, 12),
            ("Quality review & outlier handling",                              4,  0,  0,  8),
            ("New project hook (auto-extraction on repo creation via CI)",     0,  8,  8,  0),
            ("Coverage report & final handover",                               0,  0,  0,  6),
        ],
    },
]

# Datacore AI roles (monthly pricing, vision)
datacore_roles = [
    ("AI Chief of Staff", "Org-wide operational intelligence \u2014 answers any question about your repos, decisions, team activity", 800),
    ("Insight Agent", "Proactive pattern detection \u2014 surfaces what you didn\u2019t know to ask across teams", 600),
    ("Onboarding Companion", "Interactive guide for new developers \u2014 built from your actual conventions", 400),
]


def render_sheet(ws, internal=False):
    def brd(r, b=thin_border):
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).border = b

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 16

    row = 1
    ws.merge_cells(f"A1:H1")
    t = "PLUR Enterprise \u2014 Offer (INTERNAL)" if internal else "PLUR Enterprise \u2014 Offer"
    ws["A1"].value = t
    ws["A1"].font = header_font
    row = 3

    # === PARAMETERS ===
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Parameters (adjust yellow cells)").font = Font(name="Calibri", size=12, bold=True, color="E65100")
    row += 1

    params = {}
    def add_p(r, label, val, fmt=None):
        ws.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, bold=True)
        c = ws.cell(row=r, column=3, value=val)
        style_param(c)
        c.alignment = Alignment(horizontal="right")
        if fmt: c.number_format = fmt
        return f"$C${r}"

    params["seats"] = add_p(row, "Number of seats", 50, '0'); row += 1
    params["seat_price"] = add_p(row, "List price per seat (EUR/month)", 70, '#,##0'); row += 1
    params["fp_discount"] = add_p(row, "Founding Partner discount", 0.30, '0%'); row += 1
    params["rate"] = add_p(row, "Consulting rate (EUR/hour)", 85, '#,##0'); row += 1
    params["commitment"] = add_p(row, "Subscription commitment (months)", 12, '0'); row += 1
    if internal:
        params["int_rate"] = add_p(row, "Internal team rate (EUR)", 60, '#,##0'); row += 1
        params["plat_cost"] = add_p(row, "Platform internal cost (mostly built)", 3000, '#,##0'); row += 1
    row += 1

    # Team
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Team").font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for name, role in team_info:
        ws.cell(row=row, column=2, value=name).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=3, value=role).font = body_font
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=NC)
        row += 1
    row += 1

    # ================================================================
    # TIMELINE
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Timeline").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    timeline_items = [
        ("Early May", "Contract signed", ""),
        ("May W1\u20132", "Integration + first 5 repos", "GitLab SSO, deployment in parallel with scanning first 5 repos"),
        ("May W2\u20133", "Feedback loop", "Present results to tech leads, tune extraction pipeline"),
        ("May W3\u20134", "Test run + next batch", "Infra live, 10-15 users collecting memories. Scan next 10-15 repos with tuned pipeline."),
        ("June", "Onboarding month \u2014 subscription starts", "50 users, setup workshop. 20-30 repos curated, packs deployed."),
        ("June\u2013July", "Knowledge Engineering Phase A completes", "Remaining active repos, pipeline refinement"),
        ("July\u2013Aug", "Knowledge Engineering Phase B", "1,370 repos automated, new project hook, coverage report, handover"),
        ("Q3/Q4 2026", "Datacore Enterprise (scoped together)", "AI Development Team \u2014 Chief of Staff, Insight Agent, Onboarding Companion"),
    ]

    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).fill = section_fill
    ws.cell(row=row, column=2, value="When").font = section_font
    ws.cell(row=row, column=4, value="Milestone").font = section_font
    ws.cell(row=row, column=6, value="Details").font = section_font
    row += 1

    for when, milestone, details in timeline_items:
        ws.cell(row=row, column=2, value=when).font = total_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=4, value=milestone).font = body_font
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value=details).font = note_font
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=NC)
        brd(row)
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Subscription starts at go-live (June 2026), not at contract signing.").font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
    row += 1
    ws.cell(row=row, column=2, value="Integration + test run (May) included \u2014 no subscription charges before go-live.").font = Font(name="Calibri", size=11, color="2E7D32")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
    row += 2

    # ================================================================
    # 1. PLUR ENTERPRISE — ORGANIZATIONAL LEARNING
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="1. PLUR ENTERPRISE \u2014 Organizational Learning (subscription)").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    # List price
    ws.cell(row=row, column=2, value="List price per seat").font = grey_font
    c = ws.cell(row=row, column=8)
    c.value = f"={params['seat_price']}"
    c.font = grey_font
    c.alignment = Alignment(horizontal="right")
    c.number_format = '#,##0 "EUR/mo"'
    row += 1

    # Founding Partner price per seat
    dp_seat_row = row
    ws.cell(row=row, column=2, value="Founding Partner price per seat").font = discount_font
    c = ws.cell(row=row, column=8)
    c.value = f"={params['seat_price']}*(1-{params['fp_discount']})"
    c.font = discount_font
    c.alignment = Alignment(horizontal="right")
    c.number_format = '#,##0 "EUR/mo"'
    row += 1

    # Monthly
    sub_monthly_row = row
    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).fill = recurring_fill
        ws.cell(row=row, column=col).border = thick_border
    ws.cell(row=row, column=2).value = "Monthly subscription"
    ws.cell(row=row, column=2).font = recurring_font
    ws.cell(row=row, column=6, value="seats \u00d7 price =").font = grey_font
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
    c = ws.cell(row=row, column=8)
    c.value = f"={params['seats']}*H{dp_seat_row}"
    c.font = recurring_font
    c.fill = recurring_fill
    c.alignment = Alignment(horizontal="right")
    c.number_format = '#,##0 "EUR/mo"'
    row += 1

    # Annual
    sub_annual_row = row
    ws.cell(row=row, column=2, value="Annual commitment").font = total_font
    c = ws.cell(row=row, column=8)
    c.value = f"=H{sub_monthly_row}*{params['commitment']}"
    c.font = total_font
    c.alignment = Alignment(horizontal="right")
    c.number_format = eur_fmt
    row += 1

    # Savings vs list
    ws.cell(row=row, column=2, value="Annual savings vs list price").font = Font(name="Calibri", size=11, color="2E7D32")
    c = ws.cell(row=row, column=8)
    c.value = f"={params['seats']}*{params['seat_price']}*{params['fp_discount']}*{params['commitment']}"
    c.font = Font(name="Calibri", size=11, color="2E7D32")
    c.alignment = Alignment(horizontal="right")
    c.number_format = '#,##0 "EUR"'
    row += 2

    # What's included
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="What\u2019s included:").font = subsection_font
    row += 1
    for idx, item in enumerate(subscription_includes, 1):
        ws.cell(row=row, column=1, value=idx).font = body_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=item).font = body_font
        brd(row)
        row += 1

    # Market comparison
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Market context:").font = subsection_font
    row += 1
    comparisons = [
        ("GitHub Copilot Enterprise", "$39-60/seat/mo", "Code completion, 28-day memory, no learning"),
        ("Augment Code Standard", "$60/seat/mo", "Project memory, no correction-based learning"),
        ("Sourcegraph Cody Enterprise", "$59/seat/mo", "Code search, no persistent memory"),
        ("JetBrains AI Enterprise", "$60+/seat/mo", "IDE AI features, no org memory"),
        ("PLUR Enterprise (list)", "\u20ac70/seat/mo", "Persistent org learning, correction-based, knowledge graph"),
        ("PLUR Enterprise (Founding Partner)", "\u20ac49/seat/mo", "Same \u2014 30% Founding Partner discount"),
    ]
    for name, price, notes in comparisons:
        ws.cell(row=row, column=2, value=name).font = body_font
        ws.cell(row=row, column=5, value=price).font = total_font
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=notes).font = note_font
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=NC)
        brd(row)
        row += 1
    row += 2

    # ================================================================
    # 2. KNOWLEDGE ENGINEERING (one-time project)
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="2. KNOWLEDGE ENGINEERING (one-time project)").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Scan your codebase, extract conventions, build your custom ingest pipeline").font = italic_grey
    row += 1

    hdrs = ["#", "Deliverable", "Gregor", "Tadej", "Marko", "Crt", "Total h", "Cost (EUR)"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = section_font
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left" if col <= 2 else "right")
    row += 1

    ke_subtotal_rows = []

    for phase in custom_phases:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        ws.cell(row=row, column=1, value=phase["name"]).font = subsection_font
        ws.cell(row=row, column=1).fill = subsection_fill
        row += 1

        item_rows = []
        for idx, item in enumerate(phase["items"], 1):
            desc = item[0]
            phrs = list(item[1:5])
            ws.cell(row=row, column=1, value=idx).font = body_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=desc).font = body_font
            for i, ph in enumerate(phrs):
                c = ws.cell(row=row, column=3 + i)
                if ph > 0:
                    c.value = ph
                    c.font = body_font
                    style_param(c)
                else:
                    c.value = None
                c.alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=7, value=f"=SUM(C{row}:F{row})").font = body_font
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
            c = ws.cell(row=row, column=8, value=f"=G{row}*{params['rate']}")
            c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
            brd(row)
            item_rows.append(row)
            row += 1

        sub_row = row
        dash = "\u2014"
        short = phase["name"].split(dash)[0].strip()
        ws.cell(row=row, column=2, value=f"Subtotal \u2014 {short}").font = total_font
        for i in range(4):
            pl = chr(67 + i)
            c = ws.cell(row=row, column=3 + i)
            c.value = f"=SUM({pl}{item_rows[0]}:{pl}{item_rows[-1]})"
            c.font = total_font; c.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7, value=f"=SUM(G{item_rows[0]}:G{item_rows[-1]})").font = total_font
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
        c = ws.cell(row=row, column=8, value=f"=SUM(H{item_rows[0]}:H{item_rows[-1]})")
        c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row, phase_border)
        row += 1

        tok_row = row
        ws.cell(row=row, column=2, value="AI compute (tokens)").font = italic_grey
        c = ws.cell(row=row, column=8, value=phase["token_cost"])
        style_param(c); c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row)
        row += 1

        ke_subtotal_rows.append((sub_row, tok_row))
        row += 1

    # KE total
    ke_total_row = row
    brd(row, thick_border)
    ws.cell(row=row, column=2, value="KNOWLEDGE ENGINEERING TOTAL").font = grand_font
    h_refs = "+".join(f"G{sr}" for sr, _ in ke_subtotal_rows)
    ws.cell(row=row, column=7, value=f"={h_refs}").font = grand_font
    ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")
    cost_refs = "+".join(f"H{sr}+H{tr}" for sr, tr in ke_subtotal_rows)
    c = ws.cell(row=row, column=8, value=f"={cost_refs}")
    c.font = grand_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    brd(row, thick_border)
    row += 1

    ke_dp_row = ke_total_row  # no discount on custom work

    # Deliverables
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Deliverables:").font = subsection_font
    row += 1
    for d in [
        "Custom ingest pipeline tuned to your codebase (reusable, automated)",
        "30 active repos: curated knowledge packs, reviewed with your tech leads",
        "1,370 repos: auto-processed, coverage report",
        "New project hook: every new repo auto-bootstrapped on creation",
    ]:
        ws.cell(row=row, column=2, value=f"  {d}").font = body_font
        row += 1
    row += 2

    # ================================================================
    # 3. DATACORE ENTERPRISE — AI DEVELOPMENT TEAM (future)
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="3. DATACORE ENTERPRISE \u2014 AI Development Team (Founding Partnership continues)").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Autonomous AI agents powered by your organizational learning. Separate product, same partnership. Scoped together after PLUR is in place.").font = italic_grey
    row += 2

    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).fill = datacore_fill
    ws.cell(row=row, column=2, value="AI Role").font = datacore_font
    ws.cell(row=row, column=6, value="Description").font = datacore_font
    ws.cell(row=row, column=8, value="Indicative").font = datacore_font
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    row += 1

    dc_role_rows = []
    for name, desc, monthly in datacore_roles:
        for col in range(1, NC + 1):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws.cell(row=row, column=2, value=name).font = total_font
        ws.cell(row=row, column=6, value=desc).font = note_font
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        c = ws.cell(row=row, column=8, value=monthly)
        style_param(c); c.alignment = Alignment(horizontal="right"); c.number_format = '#,##0 "EUR/mo"'
        brd(row)
        dc_role_rows.append(row)
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Pricing model: monthly per AI role. Exact scope determined together after Phase 1+2.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
    row += 1
    ws.cell(row=row, column=2, value="Founding Partnership discount applies to Datacore on same terms as PLUR.").font = note_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
    row += 3

    # ================================================================
    # FIRST-YEAR COMMITMENT
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="FIRST-YEAR COMMITMENT").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    a1_row = row
    ws.cell(row=row, column=2, value="PLUR Enterprise subscription (12 months, Founding Partner rate)").font = body_font
    c = ws.cell(row=row, column=8, value=f"=H{sub_annual_row}")
    c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    brd(row); row += 1

    a2_row = row
    ws.cell(row=row, column=2, value="Knowledge Engineering (custom project, standard rate)").font = body_font
    c = ws.cell(row=row, column=8, value=f"=H{ke_dp_row}")
    c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    brd(row); row += 1

    fy_row = row
    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).border = thick_border
        ws.cell(row=row, column=col).fill = annual_fill
    ws.cell(row=row, column=2, value="FIRST-YEAR TOTAL").font = annual_font
    c = ws.cell(row=row, column=8, value=f"=H{a1_row}+H{a2_row}")
    c.font = annual_font; c.fill = annual_fill
    c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    row += 1

    ws.cell(row=row, column=2, value="Monthly equivalent").font = grey_font
    c = ws.cell(row=row, column=8, value=f"=H{fy_row}/{params['commitment']}")
    c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = '#,##0 "EUR/mo"'
    row += 1

    # Savings line
    ws.cell(row=row, column=2, value="vs list price (no Founding Partner discount)").font = grey_font
    list_formula = f"={params['seats']}*{params['seat_price']}*{params['commitment']}+H{ke_total_row}"
    c = ws.cell(row=row, column=8, value=list_formula)
    c.font = grey_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    row += 1

    ws.cell(row=row, column=2, value="Your savings as Founding Partner").font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    c = ws.cell(row=row, column=8, value=f"=H{row-1}-H{fy_row}")
    c.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
    c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    row += 3

    # ================================================================
    # ROI (client sheet only — not internal detail)
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="ROI ESTIMATE").font = Font(name="Calibri", size=13, bold=True)
    row += 1

    roi_items = [
        ("Developer time searching for information (McKinsey)", "1.8 hours/day"),
        ("Context switching cost per developer/year", "~\u20ac50,000"),
        ("Saving just 15 min/day per developer (conservative)", ""),
        ("  \u2192 Annual recovered value (50 devs \u00d7 15 min \u00d7 220 days \u00d7 \u20ac85/h)", ""),
        ("  \u2192 ROI vs first-year investment", ""),
    ]
    # Row for 15 min calc
    ws.cell(row=row, column=2, value=roi_items[0][0]).font = body_font
    ws.cell(row=row, column=8, value=roi_items[0][1]).font = total_font
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    brd(row); row += 1
    ws.cell(row=row, column=2, value=roi_items[1][0]).font = body_font
    ws.cell(row=row, column=8, value=roi_items[1][1]).font = total_font
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    brd(row); row += 1
    row += 1

    ws.cell(row=row, column=2, value="Conservative: saving 15 min/day per developer").font = total_font
    brd(row); row += 1

    roi_val_row = row
    ws.cell(row=row, column=2, value="Annual recovered value").font = total_font
    # 50 devs * 0.25h * 220 days * 85 EUR
    c = ws.cell(row=row, column=8)
    c.value = f"={params['seats']}*0.25*220*{params['rate']}"
    c.font = grand_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
    brd(row); row += 1

    roi_row = row
    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).fill = annual_fill
        ws.cell(row=row, column=col).border = thick_border
    ws.cell(row=row, column=2, value="ROI (recovered value / investment)").font = annual_font
    c = ws.cell(row=row, column=8, value=f"=H{roi_val_row}/H{fy_row}")
    c.font = annual_font; c.fill = annual_fill
    c.alignment = Alignment(horizontal="right"); c.number_format = '0.0"x"'
    row += 1

    ws.cell(row=row, column=2, value="Break-even (months)").font = grey_font
    c = ws.cell(row=row, column=8)
    c.value = f"=H{fy_row}/(H{roi_val_row}/12)"
    c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = '0.0 "months"'
    row += 3

    # ================================================================
    # INTERNAL MARGIN
    # ================================================================
    if internal:
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value="MARGIN ANALYSIS (INTERNAL ONLY)").font = margin_font
        row += 1

        r1 = row
        ws.cell(row=row, column=2, value="Subscription revenue (12 months)").font = body_font
        c = ws.cell(row=row, column=8, value=f"=H{sub_annual_row}")
        c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row); row += 1

        r2 = row
        ws.cell(row=row, column=2, value="Subscription cost (platform + ~8h/mo support)").font = body_font
        c = ws.cell(row=row, column=8, value=f"={params['plat_cost']}+8*{params['int_rate']}*{params['commitment']}")
        c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row); row += 1

        r3 = row
        ws.cell(row=row, column=2, value="Subscription margin").font = total_font
        c = ws.cell(row=row, column=8, value=f"=H{r1}-H{r2}")
        c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row, phase_border); row += 1
        row += 1

        r4 = row
        ws.cell(row=row, column=2, value="Knowledge Engineering revenue").font = body_font
        c = ws.cell(row=row, column=8, value=f"=H{ke_dp_row}")
        c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row); row += 1

        r5 = row
        ws.cell(row=row, column=2, value="Knowledge Engineering cost (hours \u00d7 internal rate)").font = body_font
        c = ws.cell(row=row, column=8, value=f"=G{ke_total_row}*{params['int_rate']}")
        c.font = body_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row); row += 1

        r6 = row
        ws.cell(row=row, column=2, value="Knowledge Engineering margin").font = total_font
        c = ws.cell(row=row, column=8, value=f"=H{r4}-H{r5}")
        c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        brd(row, phase_border); row += 1
        row += 1

        mg_row = row
        for col in range(1, NC + 1):
            ws.cell(row=row, column=col).fill = margin_fill
            ws.cell(row=row, column=col).border = thick_border
        ws.cell(row=row, column=2, value="TOTAL FIRST-YEAR MARGIN").font = margin_font
        c = ws.cell(row=row, column=8, value=f"=H{r3}+H{r6}")
        c.font = margin_font; c.fill = margin_fill
        c.alignment = Alignment(horizontal="right"); c.number_format = eur_fmt
        row += 1

        ws.cell(row=row, column=2, value="Margin % on revenue").font = grey_font
        c = ws.cell(row=row, column=8, value=f"=H{mg_row}/H{fy_row}")
        c.font = total_font; c.alignment = Alignment(horizontal="right"); c.number_format = '0%'
        row += 3

    # ================================================================
    # DESIGN PARTNER BENEFITS
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Design Partner Benefits").font = Font(name="Calibri", size=12, bold=True)
    row += 1
    for b in [
        "30% Founding Partner discount on PLUR Enterprise subscription",
        "Guaranteed best rate \u2014 you will never pay more than any future customer",
        "Influence on product roadmap \u2014 your requirements built first",
        "White-label & reseller rights (pre-negotiated for future)",
        "Partnership continues with Datacore Enterprise on same terms",
        "Case study & reference customer agreement",
        "Integration phase included \u2014 no charges before go-live",
        "Weekly check-ins + quarterly strategic reviews",
    ]:
        ws.cell(row=row, column=2, value=f"  {b}").font = body_font
        row += 1
    row += 2

    # ================================================================
    # NOTES
    # ================================================================
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Notes").font = Font(name="Calibri", size=12, bold=True)
    row += 1
    notes = [
        "All prices exclude VAT.",
        "Subscription: 12-month commitment starting at go-live (June 2026). No charges during integration.",
        "Founding Partner rate: guaranteed best rate \u2014 you will never pay more than any future customer.",
        "Knowledge Engineering: invoiced on actuals at standard consulting rate, with weekly reporting.",
        "Datacore Enterprise (AI Development Team): scoped and priced as monthly AI roles after PLUR is in place.",
        "GitHub provider available as future add-on when needed.",
        "We use Claude Code and AI-assisted development \u2014 this is how we deliver fast.",
        "Yellow cells are adjustable \u2014 all totals recalculate automatically.",
        "ROI based on McKinsey research (1.8h/day searching) and GitHub/Accenture productivity studies (25-55% gain).",
    ]
    if internal:
        notes.append("INTERNAL SHEET \u2014 do not share with client.")
    for n in notes:
        ws.cell(row=row, column=2, value=f"  {n}").font = note_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NC)
        row += 1

    ws.print_area = f"A1:H{row}"


# Generate
ws_client = wb.active
ws_client.title = "Client"
render_sheet(ws_client, internal=False)

ws_internal = wb.create_sheet("Internal")
render_sheet(ws_internal, internal=True)

out = "/Users/gregor/Data/5-plur/2-projects/plur/docs/enterprise/PLUR-Enterprise-Cost-Breakdown.xlsx"
wb.save(out)
print(f"Saved: {out}")
print()
print("1. PLUR Enterprise: 50 seats x EUR 70 list, EUR 49 Founding Partner = EUR 2,450/mo")
print("2. Knowledge Engineering: hourly custom work, -20% Founding Partner")
print("3. Datacore Enterprise: AI roles, monthly pricing, vision (scoped later)")
print()
print("Reframed: Organizational Learning, not Shared Memory")
print("Two products: PLUR (learning) + Datacore (orchestration)")
